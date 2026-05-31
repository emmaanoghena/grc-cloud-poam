import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime, timedelta
import random

# ── Scoring maps ──────────────────────────────────────────────────────────────
SEVERITY_W   = {"Critical": 10, "High": 7, "Medium": 4, "Low": 2}
ASSET_CRIT_W = {"High": 3, "Medium": 2, "Low": 1}
EXPOSURE_W   = {"Internet": 3, "Internal": 2, "Isolated": 1}
DATA_SENS_W  = {"PII": 3, "Confidential": 2, "Public": 1}

DUE_DAYS = {"Critical": 15, "High": 30, "Medium": 60, "Low": 90}

def risk_score(row):
    s = SEVERITY_W.get(row["severity"], 2)
    a = ASSET_CRIT_W.get(row["asset_criticality"], 1)
    e = EXPOSURE_W.get(row["exposure"], 1)
    d = DATA_SENS_W.get(row["data_sensitivity"], 1)
    return round(((s * 4) + (a * 3) + (e * 2) + (d * 1)) * 2.5, 1)

def risk_label(score):
    if score >= 80: return "Critical"
    if score >= 60: return "High"
    if score >= 40: return "Medium"
    return "Low"

# ── Load & enrich data ────────────────────────────────────────────────────────
df = pd.read_csv("data/mock_findings.csv")
today = datetime(2026, 5, 31)

df["risk_score"]  = df.apply(risk_score, axis=1)
df["risk_level"]  = df["risk_score"].apply(risk_label)
df["due_date"]    = df["severity"].apply(
    lambda s: (today + timedelta(days=DUE_DAYS.get(s, 90))).strftime("%Y-%m-%d")
)
df = df.sort_values("risk_score", ascending=False).reset_index(drop=True)

COLS = [
    "finding_id", "source", "service", "description", "severity",
    "risk_score", "asset", "data_sensitivity", "exposure", "owner",
    "due_date", "status", "remediation", "control_id"
]
HEADERS = [
    "Finding ID", "Source", "Service", "Description", "Severity",
    "Risk Score", "Asset", "Data Class", "Exposure", "Owner",
    "Due Date", "Status", "Remediation", "Control ID"
]
df = df[COLS]

# ── Color palette ─────────────────────────────────────────────────────────────
COLORS = {
    "critical_fill": "C0392B", "critical_font": "FFFFFF",
    "high_fill":     "E67E22", "high_font":     "FFFFFF",
    "medium_fill":   "F1C40F", "medium_font":   "2C3E50",
    "low_fill":      "27AE60", "low_font":      "FFFFFF",
    "header_bg":     "1A252F", "header_font":   "ECF0F1",
    "exec_bg":       "2C3E50", "exec_font":     "ECF0F1",
    "alt_row":       "F4F6F7", "white":         "FFFFFF",
    "accent":        "2980B9", "border":        "BDC3C7",
}

RISK_FILLS = {
    "Critical": (COLORS["critical_fill"], COLORS["critical_font"]),
    "High":     (COLORS["high_fill"],     COLORS["high_font"]),
    "Medium":   (COLORS["medium_fill"],   COLORS["medium_font"]),
    "Low":      (COLORS["low_fill"],      COLORS["low_font"]),
}

STATUS_FILLS = {
    "Open":        ("E74C3C", "FFFFFF"),
    "In Progress": ("F39C12", "FFFFFF"),
    "Closed":      ("27AE60", "FFFFFF"),
}

def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom_border():
    thin = Side(style="thin",   color=COLORS["border"])
    med  = Side(style="medium", color=COLORS["header_bg"])
    return Border(left=thin, right=thin, top=thin, bottom=med)

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════
# SHEET 1 – Risk Register
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Risk Register")

total      = len(df)
n_crit     = (df["severity"]   == "Critical").sum()
n_high     = (df["severity"]   == "High").sum()
n_med      = (df["severity"]   == "Medium").sum()
n_low      = (df["severity"]   == "Low").sum()
n_open     = (df["status"]     == "Open").sum()
n_inprog   = (df["status"]     == "In Progress").sum()
avg_score  = round(df["risk_score"].mean(), 1)
max_score  = df["risk_score"].max()

ws.merge_cells("A1:N1")
title_cell = ws["A1"]
title_cell.value = "☁  Cloud Security Risk Register & POA&M"
title_cell.font  = Font(name="Arial", bold=True, size=18, color=COLORS["header_font"])
title_cell.fill  = make_fill(COLORS["header_bg"])
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:N2")
sub = ws["A2"]
sub.value = f"Generated: {today.strftime('%B %d, %Y')}   |   Classification: CONFIDENTIAL   |   Framework: NIST SP 800-53"
sub.font  = Font(name="Arial", size=10, color="95A5A6")
sub.fill  = make_fill(COLORS["header_bg"])
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

ws.row_dimensions[3].height = 14
for col in range(1, 15):
    ws.cell(3, col).fill = make_fill(COLORS["exec_bg"])

kpi_data = [
    ("A", "B", "Total Findings", total),
    ("C", "D", "🔴  Critical",    n_crit),
    ("E", "F", "🟠  High",        n_high),
    ("G", "H", "🟡  Medium",      n_med),
    ("I", "J", "🟢  Low",         n_low),
    ("K", "L", "Open Items",      n_open),
    ("M", "N", "Avg Risk Score",  avg_score),
]

for start_col, end_col, label, value in kpi_data:
    ws.merge_cells(f"{start_col}4:{end_col}4")
    ws.merge_cells(f"{start_col}5:{end_col}5")
    lbl_cell = ws[f"{start_col}4"]
    lbl_cell.value = label
    lbl_cell.font  = Font(name="Arial", size=9, color="95A5A6")
    lbl_cell.fill  = make_fill(COLORS["exec_bg"])
    lbl_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 16

    val_cell = ws[f"{start_col}5"]
    val_cell.value = value
    val_cell.font  = Font(name="Arial", bold=True, size=20, color="ECF0F1")
    val_cell.fill  = make_fill(COLORS["exec_bg"])
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 32

ws.row_dimensions[6].height = 8
for col in range(1, 15):
    ws.cell(6, col).fill = make_fill(COLORS["exec_bg"])

HEADER_ROW = 7
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(HEADER_ROW, col_idx)
    cell.value = header
    cell.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
    cell.fill      = make_fill(COLORS["accent"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thick_bottom_border()
ws.row_dimensions[HEADER_ROW].height = 28

DATA_START = 8
for r_idx, row in df.iterrows():
    excel_row  = DATA_START + r_idx
    row_values = [
        row["finding_id"], row["source"],    row["service"],
        row["description"], row["severity"], row["risk_score"],
        row["asset"],       row["data_sensitivity"], row["exposure"],
        row["owner"],       row["due_date"], row["status"],
        row["remediation"], row["control_id"],
    ]

    is_alt = (r_idx % 2 == 1)
    base_bg = COLORS["alt_row"] if is_alt else COLORS["white"]

    for c_idx, val in enumerate(row_values, 1):
        cell = ws.cell(excel_row, c_idx)
        cell.value     = val
        cell.font      = Font(name="Arial", size=9)
        cell.fill      = make_fill(base_bg)
        cell.border    = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=(c_idx in (4, 13)))

    sev   = row["severity"]
    sfill, sfont = RISK_FILLS.get(sev, ("FFFFFF", "000000"))
    c5 = ws.cell(excel_row, 5)
    c5.fill  = make_fill(sfill)
    c5.font  = Font(name="Arial", bold=True, size=9, color=sfont)
    c5.alignment = Alignment(horizontal="center", vertical="center")

    score = row["risk_score"]
    rl    = risk_label(score)
    rfill, rfont = RISK_FILLS.get(rl, ("FFFFFF", "000000"))
    c6 = ws.cell(excel_row, 6)
    c6.fill  = make_fill(rfill)
    c6.font  = Font(name="Arial", bold=True, size=10, color=rfont)
    c6.alignment = Alignment(horizontal="center", vertical="center")

    status = row["status"]
    stfill, stfont = STATUS_FILLS.get(status, ("FFFFFF", "000000"))
    c12 = ws.cell(excel_row, 12)
    c12.fill  = make_fill(stfill)
    c12.font  = Font(name="Arial", bold=True, size=9, color=stfont)
    c12.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[excel_row].height = 40

col_widths = [12, 18, 12, 46, 11, 11, 24, 14, 12, 22, 13, 13, 46, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A8"

# ════════════════════════════════════════════════════════════════
# SHEET 2 – Executive Summary
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Executive Summary")

def exec_header(ws, text, row, span="A:H"):
    ws.merge_cells(f"{span.split(':')[0]}{row}:{span.split(':')[1]}{row}")
    c = ws[f"{span.split(':')[0]}{row}"]
    c.value     = text
    c.font      = Font(name="Arial", bold=True, size=12, color=COLORS["header_font"])
    c.fill      = make_fill(COLORS["header_bg"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    return c

def write_kv(ws, row, key, val, key_bg="F0F3F4"):
    c1 = ws.cell(row, 1, key)
    c1.font      = Font(name="Arial", bold=True, size=10)
    c1.fill      = make_fill(key_bg)
    c1.alignment = Alignment(vertical="center", indent=1)
    c1.border    = thin_border()
    ws.merge_cells(f"B{row}:H{row}")
    c2 = ws.cell(row, 2, val)
    c2.font      = Font(name="Arial", size=10)
    c2.fill      = make_fill(COLORS["white"])
    c2.alignment = Alignment(vertical="center", indent=1)
    c2.border    = thin_border()
    ws.row_dimensions[row].height = 20

ws2.merge_cells("A1:H1")
t = ws2["A1"]
t.value     = "Executive Summary – Cloud Security Posture"
t.font      = Font(name="Arial", bold=True, size=16, color=COLORS["header_font"])
t.fill      = make_fill(COLORS["header_bg"])
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 38

ws2.merge_cells("A2:H2")
d = ws2["A2"]
d.value     = f"Reporting Period: {today.strftime('%B %Y')}   |   Prepared by: GRC Automation Script"
d.font      = Font(name="Arial", size=10, color="7F8C8D")
d.fill      = make_fill(COLORS["exec_bg"])
d.alignment = Alignment(horizontal="center")
ws2.row_dimensions[2].height = 18

exec_header(ws2, "A.  PROGRAM OVERVIEW", 4)
kv_pairs_a = [
    ("Total Findings",         total),
    ("Average Risk Score",     avg_score),
    ("Highest Risk Score",     max_score),
    ("Open Items",             n_open),
    ("In Progress",            n_inprog),
    ("Report Generated",       today.strftime("%Y-%m-%d")),
]
for i, (k, v) in enumerate(kv_pairs_a, 5):
    write_kv(ws2, i, k, v)

exec_header(ws2, "B.  SEVERITY BREAKDOWN", 12)
sev_headers = ["Severity", "Count", "% of Total", "Avg Risk Score", "Open", "In Progress"]
for ci, h in enumerate(sev_headers, 1):
    c = ws2.cell(13, ci)
    c.value     = h
    c.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
    c.fill      = make_fill(COLORS["accent"])
    c.alignment = Alignment(horizontal="center")
    c.border    = thin_border()
ws2.row_dimensions[13].height = 22

for row_off, sev in enumerate(["Critical", "High", "Medium", "Low"], 14):
    sub = df[df["severity"] == sev]
    cnt = len(sub)
    pct = f"{cnt/total*100:.0f}%" if total else "0%"
    avg = round(sub["risk_score"].mean(), 1) if cnt else 0
    op  = (sub["status"] == "Open").sum()
    inp = (sub["status"] == "In Progress").sum()

    sfill, sfont = RISK_FILLS.get(sev, ("FFFFFF", "000000"))
    vals = [sev, cnt, pct, avg, op, inp]
    for ci, val in enumerate(vals, 1):
        c = ws2.cell(row_off, ci)
        c.value     = val
        c.font      = Font(name="Arial", bold=(ci == 1), size=10,
                           color=sfont if ci == 1 else "2C3E50")
        c.fill      = make_fill(sfill) if ci == 1 else make_fill(COLORS["alt_row"] if row_off % 2 == 0 else COLORS["white"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws2.row_dimensions[row_off].height = 20

exec_header(ws2, "C.  TOP 5 HIGHEST RISK FINDINGS", 19)
top5_heads = ["Finding ID", "Description", "Severity", "Risk Score", "Owner", "Due Date"]
for ci, h in enumerate(top5_heads, 1):
    c = ws2.cell(20, ci)
    c.value     = h
    c.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
    c.fill      = make_fill(COLORS["accent"])
    c.alignment = Alignment(horizontal="center")
    c.border    = thin_border()
ws2.row_dimensions[20].height = 22

for i, (_, row) in enumerate(df.head(5).iterrows(), 21):
    sev   = row["severity"]
    sfill, sfont = RISK_FILLS.get(sev, ("FFFFFF", "000000"))
    vals = [
        row["finding_id"], row["description"], sev,
        row["risk_score"], row["owner"], row["due_date"]
    ]
    for ci, val in enumerate(vals, 1):
        c = ws2.cell(i, ci)
        c.value     = val
        c.font      = Font(name="Arial", size=9,
                           bold=(ci == 3), color=sfont if ci == 3 else "2C3E50")
        c.fill      = make_fill(sfill) if ci == 3 else make_fill(COLORS["alt_row"] if i % 2 == 0 else COLORS["white"])
        c.alignment = Alignment(vertical="center", wrap_text=(ci == 2))
        c.border    = thin_border()
    ws2.row_dimensions[i].height = 38

exec_header(ws2, "D.  REMEDIATION TIMELINE (Days to Remediate)", 27)
for ci, h in enumerate(["Severity", "Target SLA (Days)", "Earliest Due Date", "Findings Due This Month"], 1):
    c = ws2.cell(28, ci)
    c.value     = h
    c.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
    c.fill      = make_fill(COLORS["accent"])
    c.alignment = Alignment(horizontal="center")
    c.border    = thin_border()
ws2.row_dimensions[28].height = 22

for row_off, (sev, days) in enumerate(DUE_DAYS.items(), 29):
    sub  = df[df["severity"] == sev]
    earliest = sub["due_date"].min() if len(sub) else "N/A"
    this_mo  = sub[sub["due_date"].str.startswith(today.strftime("%Y-%m"))]["finding_id"].count()
    sfill, sfont = RISK_FILLS.get(sev, ("FFFFFF", "000000"))
    for ci, val in enumerate([sev, days, earliest, this_mo], 1):
        c = ws2.cell(row_off, ci)
        c.value     = val
        c.font      = Font(name="Arial", size=10, bold=(ci == 1),
                           color=sfont if ci == 1 else "2C3E50")
        c.fill      = make_fill(sfill) if ci == 1 else make_fill(COLORS["white"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws2.row_dimensions[row_off].height = 20

for col, width in zip("ABCDEFGH", [14, 42, 12, 14, 22, 14, 12, 12]):
    ws2.column_dimensions[col].width = width

# ════════════════════════════════════════════════════════════════
# SHEET 3 – Scoring Model Reference
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Scoring Model")

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value     = "Risk Scoring Model Reference"
t3.font      = Font(name="Arial", bold=True, size=14, color=COLORS["header_font"])
t3.fill      = make_fill(COLORS["header_bg"])
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

formula_text = (
    "risk_score = ( (severity_weight × 4) + "
    "(asset_criticality × 3) + "
    "(exposure × 2) + "
    "(data_sensitivity × 1) ) × 2.5"
)
ws3.merge_cells("A2:E2")
f2 = ws3["A2"]
f2.value     = formula_text
f2.font      = Font(name="Courier New", bold=True, size=10, color="27AE60")
f2.fill      = make_fill("1A252F")
f2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[2].height = 28

tables = [
    ("Severity", "Weight", [("Critical", 10), ("High", 7), ("Medium", 4), ("Low", 2)],
     5, "A", RISK_FILLS),
    ("Asset Criticality", "Weight", [("High", 3), ("Medium", 2), ("Low", 1)],
     5, "C", {"High": ("C0392B", "FFFFFF"), "Medium": ("E67E22", "FFFFFF"), "Low": ("27AE60", "FFFFFF")}),
    ("Exposure", "Weight", [("Internet", 3), ("Internal", 2), ("Isolated", 1)],
     5, "E", {"Internet": ("C0392B", "FFFFFF"), "Internal": ("E67E22", "FFFFFF"), "Isolated": ("27AE60", "FFFFFF")}),
]

for title, sub_title, rows, start_row, col, fills in tables:
    c = ws3.cell(start_row - 1, ord(col) - 64)
    ws3.merge_cells(f"{col}{start_row-1}:{col}{start_row-1}")
    c.value     = f"{title} / {sub_title}"
    c.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
    c.fill      = make_fill(COLORS["accent"])
    c.alignment = Alignment(horizontal="center")
    c.border    = thin_border()
    ws3.row_dimensions[start_row - 1].height = 20

    c2 = ws3.cell(start_row - 1, ord(col) - 64 + 1)
    c2.fill   = make_fill(COLORS["accent"])
    c2.border = thin_border()

    for r_off, (label, weight) in enumerate(rows, start_row):
        fill_hex, font_hex = fills.get(label, ("FFFFFF", "000000"))
        ca = ws3.cell(r_off, ord(col) - 64)
        ca.value     = label
        ca.font      = Font(name="Arial", size=10, bold=True, color=font_hex)
        ca.fill      = make_fill(fill_hex)
        ca.alignment = Alignment(horizontal="center")
        ca.border    = thin_border()

        cb = ws3.cell(r_off, ord(col) - 64 + 1)
        cb.value     = weight
        cb.font      = Font(name="Arial", size=10)
        cb.fill      = make_fill(COLORS["white"])
        cb.alignment = Alignment(horizontal="center")
        cb.border    = thin_border()
        ws3.row_dimensions[r_off].height = 20

range_start = 11
ws3.merge_cells(f"A{range_start}:E{range_start}")
rh = ws3[f"A{range_start}"]
rh.value     = "Risk Score Ranges"
rh.font      = Font(name="Arial", bold=True, size=10, color=COLORS["header_font"])
rh.fill      = make_fill(COLORS["accent"])
rh.alignment = Alignment(horizontal="center")
rh.border    = thin_border()
ws3.row_dimensions[range_start].height = 20

for r_off, (label, rng) in enumerate([
    ("Critical", "80 – 100"),
    ("High",     "60 – 79"),
    ("Medium",   "40 – 59"),
    ("Low",      "0 – 39"),
], range_start + 1):
    fill_hex, font_hex = RISK_FILLS.get(label, ("FFFFFF", "000000"))
    ws3.merge_cells(f"A{r_off}:B{r_off}")
    ca = ws3[f"A{r_off}"]
    ca.value     = label
    ca.font      = Font(name="Arial", bold=True, size=10, color=font_hex)
    ca.fill      = make_fill(fill_hex)
    ca.alignment = Alignment(horizontal="center")
    ca.border    = thin_border()

    ws3.merge_cells(f"C{r_off}:E{r_off}")
    cb = ws3[f"C{r_off}"]
    cb.value     = rng
    cb.font      = Font(name="Arial", size=10)
    cb.fill      = make_fill(COLORS["white"])
    cb.alignment = Alignment(horizontal="center")
    cb.border    = thin_border()
    ws3.row_dimensions[r_off].height = 20

for col, width in zip("ABCDE", [20, 10, 20, 10, 20]):
    ws3.column_dimensions[col].width = width

# ── Save ──────────────────────────────────────────────────────────────────────
wb._sheets = [ws2, ws, ws3]

out_path = "output/cloud_security_poam.xlsx"
wb.save(out_path)
print(f"Workbook saved → {out_path}")
print(f"Findings: {total} | Avg Risk: {avg_score} | Max Risk: {max_score}")
